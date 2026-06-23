import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from clienteApp.models import Venta

def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas Luppy"

    # Estilos personalizados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Verde Luppy
    center_alignment = Alignment(horizontal="center")

    # Encabezados
    headers = ['Fecha y Hora', 'Cliente', 'Total Venta', 'Estado']
    ws.append(headers)

    # Aplicar estilos a los encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Datos
    ventas = Venta.objects.all().order_by('-fecha_venta')

    for v in ventas:
        ws.append([
            v.fecha_venta.replace(tzinfo=None).strftime("%d/%m/%Y %H:%M"),
            str(v.id_cliente),  # Asumiendo que quieres ver el nombre del cliente
            float(v.total),
            v.get_estado_display() if hasattr(v, 'get_estado_display') else v.estado
        ])

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas_Luppy.xlsx"'
    wb.save(response)
    return response